from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import cell
import datetime

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("/home/urik/Python38/workspace/sample.xlsx")

LoadedXLS = load_workbook ('/media/urik/Windows7_OS/temp/Operative Emails.xlsx')
print(LoadedXLS.properties)
for sheetName in LoadedXLS.sheetnames:
    print(sheetName)
    for sh in  LoadedXLS._sheets:
        print (type(sh))
        shValues = []
        for xxx in sh.values:
            #print (xxx)
            shValues.append (xxx)
        print (shValues)
pip install gmail

